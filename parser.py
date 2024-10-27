from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
import time
import openpyxl

bn = '\n'   #amvera жалуется на \ в f-строке, поэтому так

def parser(bot, way_to_data):
    # хз, что это, но с ним устанавливается chrome
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # Запуск в фоновом режиме
    chrome_options.add_argument("--no-sandbox")  # Отключение песочницы
    chrome_options.add_argument("--disable-dev-shm-usage")  # Использование временной папки для файлов

    service = Service("/usr/bin/chromedriver")  # Укажите путь к драйверу
    driver = webdriver.Chrome(service=service, options=chrome_options)
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--window-size=1920x1080")

    driver.set_window_position(0, 0)
    driver.set_window_size(1920, 1080)

    options = Options()

    options.add_argument("start-maximized")  # Открыть браузер в максимизированном режиме
    options.add_argument("disable-infobars")  # Отключить информационные панели
    options.add_argument("--disable-extensions")  # Отключить расширения
    options.add_argument("--disable-gpu")  # Применимо только для Windows
    options.add_argument("--disable-dev-shm-usage")  # Решение проблемы ограниченных ресурсов
    options.add_argument("--no-sandbox")  # Обойти модель безопасности ОС
    #хз, что это, но с ним устанавливается chrome


    driver.get("https://cabinet.miem.hse.ru/#/projects/sandbox")  # получаем код страницы "песочницы"
    time.sleep(2)

    main_page = driver.page_source  #получаем html код

    book = openpyxl.open(way_to_data)
    sheet = book['settings']        #открываем страницу настроек
    book.active = book['settings']
    projects_list = sheet[0 + 1][3 + 0].value.split('_')    #получаем список уже просмотренных проектов


    look_list = []  #сохраняем номера проектов и потом посмотрим их
    list_counter = 0  # будем считать кол-во проектов, которые есть в projects_list, если таких проектов будет 5,
                        # то значит мы уже точно смотрим старые проекты


    def finder(start, end, shift_start, shift_end, main_page): #функция, которая ищет в html-коде объекты по начальной
                                                                # строке и по последней строке
        start_index = main_page.find(start) + shift_start
        main_page = main_page[start_index:] #обрезаем сначала по слову start
        end_index = main_page.find(end) + shift_end
        return main_page[0: end_index], main_page   #возвращаем искомый объект и обрезанный с начала html код


    while list_counter < 5: #пока не нашли 5 просмотренных проекта
        project_number, main_page = finder('#/project/', '/', 10, 0, main_page)#получаем номер проекта
        if project_number in projects_list:  #если такой номер уже есть, то увеличиваем счетчик
            list_counter += 1
        else:
            look_list.append(project_number)

    print("Новые проекты:", look_list)  #полезная информация в консоль

    for i in look_list: #просматриваем все новые проекты
        driver.get(f"https://cabinet.miem.hse.ru/#/project/{i}/")
        time.sleep(2)

        main_page = driver.page_source  #получаем html код страницы проекта


        #тип проекта
        project_type, main_page = finder('Иконка типа проекта', '>', 0, 0, main_page)
        project_type, main_page = finder('<div data-', '<', 25, 0, main_page)

        #название
        project_name, main_page = finder('project-header__main-info--primary', '</div>', 37, 0, main_page)

        #цель
        project_goal, main_page = finder('<p data-nodeid=', '</div>', 0, 0, main_page)

        #ожидаемые результаты
        project_results, main_page = finder('Ожидаемые результаты', '>', 0, 0, main_page)
        main_page = main_page.replace('</div>', '', 1)
        project_results, main_page = finder('<p data-nodeid=', '</div>', 0, 0, main_page)

        #требуемые и приобретаемые навыки
        temp, main_page = finder('semibold mb-1 information-card-competence-title', '<', 0, 0, main_page)
        main_page = main_page.replace('</div>', '', 1) #нужно убрать ближайший </div>, который еще до требований
        project_requirements, main_page = finder('<p data-nodeid=', '</div>', 0, 0, main_page)


        #переходим на страницу с ВАКАНСИЯМИ
        driver.get(f"https://cabinet.miem.hse.ru/#/project/{i}/vacancies")
        time.sleep(2)

        main_page = driver.page_source  #получаем html код страницы с вакансиями

        all_positions = ''  #строка с информацией о вакансия
        while "hidden-xs-only list__row list__row--non-clickable el-row el-row--flex" in main_page:
            #находим блок с вакансиями
            temp, main_page = finder("hidden-xs-only list__row list__row--non-clickable el-row el-row--flex", '>', 0, 0, main_page)

            #находим позицию, на которую ищут
            position, main_page = finder('<div data-v-', '</div>', 0, 0, main_page)
            main_page = main_page.replace('<div data-v-', '<', 1)   #убираем метку, на которую мы ориентировались

            #ищем кол-во человек на данную позицию
            amount, main_page = finder('<div data-v-', '</div>', 0, 0, main_page)
            main_page = main_page.replace('<div data-v-', '<', 1)

            #требования к участникам
            requirements, main_page = finder('<div data-v-', '</div></div>', 0, 0, main_page)
            main_page = main_page.replace('<div data-v-', '<', 1)

            #желаемые требования
            welcome, main_page = finder('</div></div>', '</div></div>', 6, 0, main_page)

            #обрабатываем требования, чтобы они красиво выглядели
            #   - заменяем на \n, чтобы требования не были в одну строчку
            all_positions = all_positions + (f"{bn} {bn}[б]Позиция:[/б] {position}{bn}[б]Кол-во:[/б] {amount}{bn}[б]Требования:[/б] "
                                             f"{requirements.replace('- -','-').replace('-',f'{bn}•')}{bn}[б]Желательно:[/б] "
                                             f"{welcome.replace('- -','-').replace('-',f'{bn}•')}{bn}")


        #ЧИСТИМ от html тегов;  по [splitter] будем потом обратно делить на разные типы информации
        res = f"{project_type}[splitter]{project_name}[splitter]{project_goal}[splitter]{project_results}[splitter]{project_requirements}[splitter]{all_positions}"
        while (res.find('<') < res.find('>')) and res.find('<')!=-1 and res.find('>')!=-1:
            first_ind = res.find('<')
            second_ind = res.find('>')
            res = res.replace(str(res[first_ind: second_ind + 1]), '')  #убираем всю информацию между < и >

        while '\n\n' in res:    #убираем лишние \n
            res = res.replace('\n\n','\n')


        array = res.split('[splitter]') #после очистки разделяем обратно все переменные
        #добавляем ссылки и жирность
        res = (f'<b>Проект №<a href="https://cabinet.miem.hse.ru/#/project/{i}/">{i}</a>{bn}{bn}</b><b>Тип</b>: {array[0]}{bn}{bn}'
               f'<b>Название:</b> {array[1]}{bn}{bn}<b>Цель:</b> {array[2]}{bn}{bn}<b>Ожидаемые результаты:</b>\n {array[3]}{bn}{bn}<b>'
               f'Требования:</b>{bn} {array[4]}{bn}{bn}<b>Вакансии(<a href="https://cabinet.miem.hse.ru/#/project/{i}/vacancies">'
               f'ссылка</a>):</b> {array[5]}')
        #добавляем теги жирности (такой костыль нужен, чтобы на этапе очистки теги не удалились)
        res = res.replace('[б]', "<b>").replace('[/б]', "</b>").replace('\n\n\n', '\n\n')


        print(f"Текст для рассылки:\n[{res}]")

        #делаем рассылку для всех, кто ее получает
        sheets_list = book.sheetnames  # получаем список листов

        for j in sheets_list:       #проходимся по всем листам
            if j == 'settings':
                continue
            else:
                sheet = book[j]
                book.active = book[j]

            if  sheet.max_column > 4 and sheet[0 + 1][4 + 0].value: #если рассылка включена, то отправляем
                print(res[:4090 + 1])
                try:
                    bot.send_message(j, res[:4090 + 1], parse_mode='HTML')
                except Exception:
                    try:
                        bot.send_message(j, res[:4090 + 1]+'>', parse_mode='HTML')
                    except Exception:
                        bot.send_message(j, f'[!] <b>Проект №<a href="https://cabinet.miem.hse.ru/#/project/{i}/">{i}</a></b>\nДанные недоступны из-за неизвестной ошибки', parse_mode='HTML')

        print(f'Рассылка проекта {i} завершена')

        #сохраняем еще один просмотренный проект на страницу настроек
        sheet = book['settings']
        book.active = book['settings']
        link = sheet.cell(row=0 + 1, column=3 + 1)
        link.value = i + '_' + sheet[0 + 1][3 + 0].value
        print(f"Проект {i} добавлен в просмотренные")

        book.save(way_to_data) #сохраняем проект