import openpyxl
from telebot import TeleBot, types
from datetime import date

def add_word(bot, message, data):
    bot_message  = bot.send_message(message.chat.id, "Отправьте пару слов через дефис(вот так: <code>слово1=слово2</code>)", parse_mode='HTML')
    bot.register_next_step_handler(message, add_word_2, bot, data, bot_message)


def add_word_2(message, bot, data, bot_message):
    #по стандарту ищем лист с id человека
    book = openpyxl.open("data/data.xlsx")
    sheets_list = book.sheetnames

    for i in sheets_list:
        if i == str(message.chat.id):
            sheet = book[i]
            book.active = book[i]  # задаем новую активную страницу
            break

    data = data.replace('add_word_to:', '') #убираем префикс


    i = 0
    while i < sheet.max_column + 1: #ищем столбец с колодой
        if sheet[0 + 1][i + 0].value == data:   #если текущая ячейка == имя колоды
            break
        i += 10

        if i > sheet.max_column + 1:    #если вдруг каким-то образом мы не нашли колоду
            bot.send_message(message.chat.id, "<code>Ошибка №4 в файле add_word</code>\nСообщить об ошибке @Tsygaika",
                             parse_mode='HTML')

    j = 2
    while sheet[j + 1][i + 0].value != None:    #ищем в колоде первую свободную строку
        j += 1


    words = message.text.split('=')

    if len(words) != 2: #если пользователь ввел не два слова
        markup = types.InlineKeyboardMarkup(row_width=1)
        btn = types.InlineKeyboardButton(text='Ввести еще раз', callback_data='add_word_to:' + data)
        markup.add(btn)
        bot.edit_message_text('Вы ввели слова в неверном формате', bot_message.chat.id,
                              message_id=bot_message.message_id, reply_markup=markup)
        return


    link = sheet.cell(row=j + 1, column=i + 1)  # записываем слова
    link.value = words[0]
    link = sheet.cell(row=j + 1, column=i + 1 + 1)
    link.value = words[1]
    link = sheet.cell(row=j + 1, column=i + 2 + 1)
    link.value = date.today()
    link = sheet.cell(row=j + 1, column=i + 3 + 1)
    link.value = 0  #сохраняем число интервалов

    book.save('data/data.xlsx')  # сохраняем


    markup = types.InlineKeyboardMarkup(row_width=1)
    btn1 = types.InlineKeyboardButton(text='Добавить еще пару', callback_data='add_word_to:' + data)
    btn2 = types.InlineKeyboardButton(text='Вернуться к колодам', callback_data='add_cards')
    markup.add(btn1, btn2)
    bot.send_message(message.chat.id, f"Пара слов {words[0]} - {words[1]} была добавлена в колоду {data}", reply_markup=markup)

    book.close()