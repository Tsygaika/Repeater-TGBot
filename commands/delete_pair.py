import openpyxl
from telebot import types

def delete_pair(bot, message):
    book = openpyxl.open("data/data.xlsx")
    sheets_list = book.sheetnames  # получаем список листов

    for i in sheets_list:  # ищем лист беседы
        if i == str(message.chat.id):
            sheet = book[i]
            book.active = book[i]  # задаем новую активную страницу
            break

    else:  # если листа для такой беседы нет, то создаем
        book.create_sheet(str(message.chat.id))
        book.active = book[str(message.chat.id)]
        sheet = book.active

    book.save('data/data.xlsx')  # сохраняем

    markup = types.InlineKeyboardMarkup(row_width=1)

    i = 0
    if sheet[0 + 1][0 + 0].value != None:  # если есть хоть одна колода

        while i < sheet.max_column + 1:  # создаем кнопки с названием каждой колоды
            btn = types.InlineKeyboardButton(text=sheet[0 + 1][i + 0].value,
                                             callback_data='delete:' + sheet[0 + 1][i + 0].value)
            markup.add(btn)
            i += 10

    if i == 0:  # если нет колод
        bot.send_message(message.chat.id, 'Вы не можете удалить пару слов, так как у вас не колод')

    else:
        bot.send_message(message.chat.id, 'Ваши колоды', reply_markup=markup)

    book.close()


def delete_pair_2(bot, call):
    bot_message = bot.edit_message_text('Отправьте первое слово в паре, которую нужно удалить',
                                        call.message.chat.id, message_id=call.message.message_id)

    bot.register_next_step_handler(call.message, delete_pair_3, bot, call)

def delete_pair_3(message, bot, call):
    first_word = message.text   #получаем первое слово из пары, которую нужно удалить
    call_data = call.data.replace('delete:', '')  # убираем префикс
    message = call.message


    # по стандарту ищем лист с id человека
    book = openpyxl.open("data/data.xlsx")
    sheets_list = book.sheetnames

    for i in sheets_list:
        if i == str(message.chat.id):
            sheet = book[i]
            book.active = book[i]  # задаем новую активную страницу
            break

    i = 0
    while i < sheet.max_column + 1:  # ищем столбец с колодой
        if sheet[0 + 1][i + 0].value == call_data:  # если текущая ячейка == имя колоды
            break
        i += 10

        if i > sheet.max_column + 1:  # если вдруг каким-то образом мы не нашли колоду
            bot.send_message(message.chat.id, "<code>Ошибка №10 в файле delete_pair</code>\nСообщить об ошибке @Tsygaika",
                             parse_mode='HTML')

    j = 2
    flag = 1
    while sheet[j + 1][i + 0].value != None:  # идем по списку пока не будет первая пустая строка
        if sheet[j + 1][i + 0].value.lower() == first_word.lower(): #не учитываем регистр
            flag = 0
            second_word = sheet[j + 1][i + 1 + 0].value

            while sheet[j + 1][i + 0].value != None:    #пока следующее слово не пусто
                link = sheet.cell(row=j + 1, column=i + 1)  #перезаписываем первое слово
                link.value = sheet[j + 1 + 1][i + 0].value
                link = sheet.cell(row=j + 1, column=i + 1 + 1)  #перезаписываем второе слово
                link.value = sheet[j + 1 + 1][i + 1 + 0].value
                link = sheet.cell(row=j + 1, column=i + 2 + 1)  #перезаписываем дату
                link.value = sheet[j + 1 + 1][i + 2 + 0].value
                link = sheet.cell(row=j + 1, column=i + 3 + 1)  #перезаписываем интервал
                link.value = sheet[j + 1 + 1][i + 3+ 0].value
                j+=1

            book.save('data/data.xlsx')
            bot.edit_message_text(f'Пара слов <b>{first_word} - {second_word}</b> удалена', message.chat.id,
                                  message_id=message.message_id, parse_mode='HTML')
            book.close()
            return

        j += 1

    if flag:
        bot.edit_message_text('Такой пары в колоде нет', message.chat.id, message_id=message.message_id)
        book.close()