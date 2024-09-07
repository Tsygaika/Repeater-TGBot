from telebot import TeleBot, types
import openpyxl
import math

def create_pack(bot, message):
    try:
        bot_message = bot.edit_message_text('Отправьте название колоды', message.chat.id, message_id = message.message_id)
    except Exception:
        bot_message = bot.send_message(message.chat.id, 'Отправьте название колоды')

    bot.register_next_step_handler(message, create_pack_2, bot, bot_message)

def create_pack_2(message, bot, bot_message):
    from shortcuts.packs_list import packs_list     #получаем список уже имеющихся колод
    packs_list = packs_list(bot, message)

    if message.text in packs_list:
        markup = types.InlineKeyboardMarkup(row_width=1)
        btn = types.InlineKeyboardButton(text='Изменить название', callback_data='create_pack')
        markup.add(btn)

        bot.edit_message_text('Колода с таким названием уже есть', message.chat.id, message_id=bot_message.message_id, reply_markup = markup)
        bot.delete_message(message.chat.id, message.message_id)
        return

    book = openpyxl.open("data/data.xlsx")
    sheets_list = book.sheetnames  # получаем список листов

    for i in sheets_list:  # ищем лист беседы
        if i == str(message.chat.id):
            sheet = book[i]
            book.active = book[i]  # задаем новую активную страницу
            break

    if sheet.max_column == 1:  # когда лист пустой
        i = 0

    else:   #если есть хоть одна колода
        i = math.ceil( sheet.max_column / 10 ) * 10 #с помощью округления вверх получаем номер следующего столбца

    link = sheet.cell(row = 1 + 1, column = i + 1)  #создаем новую колоду
    link.value = 'front'
    link = sheet.cell(row = 1 + 1, column = i + 1 + 1)
    link.value = 'back'
    link = sheet.cell(row = 1 + 1, column = i + 2 + 1)
    link.value = 'description'
    link = sheet.cell(row = 0 + 1, column = i + 1 + 1)
    link.value = "<-pack's name"
    link = sheet.cell(row = 0 + 1, column = i + 1)
    link.value = message.text

    book.save('data/data.xlsx')
    book.close()


    markup = types.InlineKeyboardMarkup(row_width=1)
    btn = types.InlineKeyboardButton(text='Добавить карточки', callback_data='packname:' + message.text)
    markup.add(btn)

    bot.edit_message_text(f'Колода «{message.text}» создана', bot_message.chat.id, message_id = bot_message.message_id, reply_markup = markup)