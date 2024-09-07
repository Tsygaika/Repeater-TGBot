#это копия файла watch.py, потом надо оптимизировать


import openpyxl
from telebot import types

def showdata(bot, message):     #тут мы создаем список колод
    book = openpyxl.open("data/data.xlsx")
    sheets_list = book.sheetnames  # получаем список листов

    for i in sheets_list:  # ищем лист беседы
        if i == str(message.chat.id):
            sheet = book[i]
            book.active = book[i]  # задаем новую активную страницу
            break

    else:  # если листа для такой беседы нет, то создаем
        book.create_sheet(str(message.chat.id))
        book.active = book[str(message.chat.id)]
        sheet = book.active

    book.save('data/data.xlsx')  # сохраняем


    markup = types.InlineKeyboardMarkup(row_width=1)

    i = 0
    if sheet[0 + 1][0 + 0].value != None:  # если есть хоть одна колода

        while i < sheet.max_column + 1: #создаем кнопки с названием каждой колоды
            btn = types.InlineKeyboardButton(text=sheet[0 + 1][i + 0].value,
                                             callback_data='showdata:' + sheet[0 + 1][i + 0].value)
            markup.add(btn)
            i += 10


    if i == 0:  #если нет колод
        btn = types.InlineKeyboardButton(text='+ создать колоду', callback_data='create_pack')
        markup.add(btn)
        bot.send_message(message.chat.id, 'У вас нет колод',reply_markup=markup)

    else:
        bot.send_message(message.chat.id, 'Ваши колоды', reply_markup=markup)

    book.close()


def open_pack(bot, call):
    call_data = call.data.replace('showdata:', '') #убираем префикс
    message = call .message

    # по стандарту ищем лист с id человека
    book = openpyxl.open("data/data.xlsx")
    sheets_list = book.sheetnames

    for i in sheets_list:
        if i == str(message.chat.id):
            sheet = book[i]
            book.active = book[i]  # задаем новую активную страницу
            break

    i = 0
    while i < sheet.max_column + 1:  # ищем столбец с колодой
        if sheet[0 + 1][i + 0].value == call_data:  # если текущая ячейка == имя колоды
            break
        i += 10

        if i > sheet.max_column + 1:  # если вдруг каким-то образом мы не нашли колоду
            bot.send_message(message.chat.id, "<code>Ошибка №8 в файле showdata</code>\nСообщить об ошибке @Tsygaika",
                             parse_mode='HTML')


    words_list = '' #список пар слов
    j = 2
    while sheet[j + 1][i + 0].value != None:  # идем по списку пока не будет первая пустая строка
        words_list = (words_list + '\n' + str(sheet[j + 1][i + 0].value) + ' - ' + str(sheet[j + 1][i + 1 + 0].value)
                      + ' - ' + str(sheet[j + 1][i + 2 + 0].value)) + ' - ' + str(sheet[j + 1][i + 3 + 0].value)#добавляем оба слова
        j += 1

    if j == 2:  #если колода пуста
        markup = types.InlineKeyboardMarkup(row_width=1)
        btn = types.InlineKeyboardButton(text='+ добавить слова', callback_data='packname:' + call_data)
        markup.add(btn)

        bot.edit_message_text(f'Колода {call_data} пуста', message.chat.id, message_id=message.message_id,
                              reply_markup=markup)

    else:
        bot.edit_message_text(f'Колода {call_data}: ' + words_list[:4000], message.chat.id, message_id=message.message_id)
        words_list = words_list[4000:]
        while len(words_list) >= 4000:
            bot.send_message(call.message.chat.id, words_list[:4000])
            words_list = words_list[4000:]

    book.close()