from datetime import date
import datetime

import openpyxl
from telebot import types

from math import floor

def repeat(bot, message):
    book = openpyxl.open("data/data.xlsx")
    sheets_list = book.sheetnames  # получаем список листов

    for i in sheets_list:  # ищем лист беседы
        if i == str(message.chat.id):
            sheet = book[i]
            book.active = book[i]  # задаем новую активную страницу
            break

    else:  # если листа для такой беседы нет, то создаем
        book.create_sheet(str(message.chat.id))
        book.active = book[str(message.chat.id)]
        sheet = book.active

    book.save('data/data.xlsx')  # сохраняем

    markup = types.InlineKeyboardMarkup(row_width=1)

    i = 0
    if sheet[0 + 1][0 + 0].value != None:  # если есть хоть одна колода

        while i < sheet.max_column + 1:  # создаем кнопки с названием каждой колоды
            btn = types.InlineKeyboardButton(text=sheet[0 + 1][i + 0].value,
                                             callback_data='repeat:' + sheet[0 + 1][i + 0].value)
            markup.add(btn)
            i += 10

    if i == 0:  # если нет колод
        btn = types.InlineKeyboardButton(text='+ создать колоду', callback_data='create_pack')
        markup.add(btn)
        bot.send_message(message.chat.id, 'У вас нет колод', reply_markup=markup)

    else:
        bot.send_message(message.chat.id, 'Ваши колоды', reply_markup=markup)

    book.close()


def repeat_2(bot, call):
    call_data = call.data.replace('repeat:', '')  # убираем префикс
    message = call.message

    # по стандарту ищем лист с id человека
    book = openpyxl.open("data/data.xlsx")
    sheets_list = book.sheetnames

    for i in sheets_list:
        if i == str(message.chat.id):
            sheet = book[i]
            book.active = book[i]  # задаем новую активную страницу
            break

    i = 0
    while i < sheet.max_column + 1:  # ищем столбец с колодой
        if sheet[0 + 1][i + 0].value == call_data:  # если текущая ячейка == имя колоды
            break
        i += 10

        if i > sheet.max_column + 1:  # если вдруг каким-то образом мы не нашли колоду
            bot.send_message(message.chat.id, "<code>Ошибка №7 в файле repeat</code>\nСообщить об ошибке @Tsygaika",
                             parse_mode='HTML')

    j = 2
    repeat_flag = 1
    while sheet[j + 1][i + 0].value != None:  # идем по списку пока не будет первая пустая строка
        #sheet[j + 1][i + 2 + 0].value == date.today()
        if (sheet[j + 1][i + 2 + 0].value - datetime.datetime.now()).days < 0:  #если разница в днях отрицательная, то повтор
            repeat_flag = 0 #если хоть одно слово было выведено, то меняем флаг

            markup = types.InlineKeyboardMarkup(row_width=1)
            btn = types.InlineKeyboardButton(text='Проверить', callback_data=f'check:{i}:{j}')
            markup.add(btn)

            bot.edit_message_text(f'Вспомните пару к слову\n\n<b>{sheet[j + 1][i + 0].value}</b>', message.chat.id,
                                  message_id=message.message_id, parse_mode='HTML', reply_markup=markup)
            book.close()
            return

        j += 1

    if repeat_flag:
        bot.edit_message_text('В этой колоде сегодня нечего повторять', message.chat.id, message_id=message.message_id)
    book.close()


def repeat_3(bot, call):
    message = call.message


    # по стандарту ищем лист с id человека
    book = openpyxl.open("data/data.xlsx")
    sheets_list = book.sheetnames

    for i in sheets_list:
        if i == str(message.chat.id):
            sheet = book[i]
            book.active = book[i]  # задаем новую активную страницу
            break


    call_data = call.data.split(':')
    i = int(call_data[1])
    j = int(call_data[2])

    markup = types.InlineKeyboardMarkup(row_width=3)
    btn1 = types.InlineKeyboardButton(text='Легко', callback_data=f'easy:{i}:{j}')
    btn2 = types.InlineKeyboardButton(text='Средне', callback_data=f'medium:{i}:{j}')
    btn3 = types.InlineKeyboardButton(text='Сложно', callback_data=f'hard:{i}:{j}')
    btn4 = types.InlineKeyboardButton(text='Повторить еще раз', callback_data=f'again:{i}:{j}')
    markup.add(btn1, btn2, btn3, btn4)

    bot.edit_message_text(f'Проверка\n\n<b>{sheet[j + 1][i + 0].value}</b> - <b>{sheet[j + 1][i + 1 + 0].value}</b>\n\n'
                          f'Насколько легко было вспомнить?', message.chat.id, message_id=message.message_id, parse_mode='HTML',
                          reply_markup=markup)
    book.close()


def edit(bot, call):
    message = call.message


    # по стандарту ищем лист с id человека
    book = openpyxl.open("data/data.xlsx")
    sheets_list = book.sheetnames

    for i in sheets_list:
        if i == str(message.chat.id):
            sheet = book[i]
            book.active = book[i]  # задаем новую активную страницу
            break


    call_data = call.data.split(':')    #достаем нужные данные из call.data
    react = call_data[0]
    i = int(call_data[1])
    j = int(call_data[2])


    prev_gap = sheet[j + 1][i + 3 + 0].value    #получаем предыдущий промежуток
        #в первые 3 этапа повторения кнопки(кроме again) не имеют значения
    if react == 'easy' or (sheet[j + 1][i + 3 + 0].value < 5 and (react == 'medium' or react == 'hard')):
        new_gap = 1.5 * prev_gap + 1 #считаем по формуле y=2.5x+1, но в этой формуле отсчет от начала, а у меня
                                      #относительный, поэтому я вычитаю prev_gap

    elif react == 'medium' and sheet.cell(row=j + 1, column=i + 3 + 1) > 5:
        new_gap = (prev_gap - 1) / 1.5  #получаем предыдущее значение промежутка

    elif react == 'hard' and sheet.cell(row=j + 1, column=i + 3 + 1) > 5:
        new_gap = ((prev_gap - 1) / 1.5 - 1) / 1.5  #откатываем промежуток на два значения назад

    elif react == 'again': #в зависимости от того, какую оценку получило слово, будем менять интервал
        new_gap = 0 #сбрасываем промежуток
    else:
        bot.send_message(message.chat.id, "<code>Ошибка №8 в файле repeat</code>\nСообщить об ошибке @Tsygaika",
                         parse_mode='HTML')
        return

    new_date = datetime.date.today() + datetime.timedelta(days=floor(new_gap))  #получаем следующую дату повтора
    link = sheet.cell(row=j + 1, column=i + 2 + 1)  #и сохраняем значения
    link.value = new_date
    link = sheet.cell(row=j + 1, column=i + 3 + 1)
    link.value = new_gap

    book.save('data/data.xlsx')
    book.close()
    next_word(bot, message, i ,j)

def check_for_todays_date(sheet, i): #проверяем нет ли в колоде еще сегодняшней даты(может быть, если нажать again)
    j = 2
    while sheet[j + 1][i + 0].value != None:  # идем по списку пока не будет первая пустая строка
        if (sheet[j + 1][i + 2 + 0].value - datetime.datetime.now()).days < 0:
            return j - 1 #так как next_word начинает искать слова со след., то, чтобы найти данное слово, нужно указать на пред.

        j+=1

    return -1


def next_word(bot, message, i ,j):

    # по стандарту ищем лист с id человека
    book = openpyxl.open("data/data.xlsx")
    sheets_list = book.sheetnames

    for _ in sheets_list:
        if _ == str(message.chat.id):
            sheet = book[_]
            book.active = book[_]  # задаем новую активную страницу
            break


    j+=1    #начинаем сразу со следующего слова
    repeat_flag = 1
    while sheet[j + 1][i + 0].value != None:  # идем по списку пока не будет первая пустая строка

        if (sheet[j + 1][i + 2 + 0].value - datetime.datetime.now()).days < 0:
            repeat_flag = 0  # если хоть одно слово было выведено, то меняем флаг

            markup = types.InlineKeyboardMarkup(row_width=1)
            btn = types.InlineKeyboardButton(text='Проверить', callback_data=f'check:{i}:{j}')
            markup.add(btn)

            bot.edit_message_text(f'Вспомните пару к слову\n\n<b>{sheet[j + 1][i + 0].value}</b>', message.chat.id,
                                  message_id=message.message_id, parse_mode='HTML', reply_markup=markup)
            book.close()
            return

        j += 1

    if repeat_flag:
        j = check_for_todays_date(sheet, i)     #если есть еще слово с сегодняшней датой
        if j != -1:
            next_word(bot, message, i, j)
        else:
            bot.edit_message_text('Больше карточек на сегодня нет', message.chat.id, message_id=message.message_id)

    book.close()